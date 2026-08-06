"""
Duchess Hub — moteur budget artiste (génération / lecture / réécriture des
fichiers .xlsx "1 fichier = 1 budget d'artiste, 1 feuillet = 1 projet").

Reproduit fidèlement la structure et les formules du classeur VBA d'origine
(00_Template_Budget_250626.xlsm, macro MacrosDuchess.bas) mais SANS macro :
toute la logique (ajouter/supprimer sous-poste ou dépense, recalcul des
totaux) est portée ici, côté serveur, et appelée depuis le Hub web.

Modèle de données (par feuillet / projet) :
  Catégorie (7 fixes, ordre imposé)
    -> Sous-poste ("container" avec dépenses, ou "simple" pour DIVERS)
         -> Dépense (uniquement pour les sous-postes "container")

Colonnes (identiques au fichier d'origine) :
  A Catégorie | B Axe analytique 3 (inutilisé pour l'instant) | C Dépense/label
  D Fournisseur | E BUDGET | F PRÉVISIONNEL | G RÉALISÉ
  H Prévi/Budget % | I Réalisé/Prévi % | J Réalisé/Budget %
  K Type (masquée : CAT / SP / DEP)

Règle héritée de la macro (Style_Container / MAJ_Plage_Bloc) :
  - E (BUDGET) est TOUJOURS une saisie manuelle au niveau Catégorie et
    Sous-poste (jamais calculé) — c'est le budget prévu, saisi par le
    producteur. Les dépenses individuelles n'ont pas de BUDGET propre.
  - F (PRÉVISIONNEL) et G (RÉALISÉ) remontent du bas vers le haut :
    Dépense (saisie) -> somme au Sous-poste -> somme à la Catégorie.
  - Un sous-poste "simple" (catégorie DIVERS) n'a pas de dépenses enfants :
    E/F/G sont saisis directement dessus.

Ce module n'évalue jamais les formules Excel lui-même (openpyxl ne le fait
pas) : les totaux/ratios renvoyés à l'API sont toujours recalculés en Python
à partir des valeurs de base, pour être fiables même si le fichier n'a
jamais été rouvert dans Excel. Les formules sont quand même écrites dans le
fichier (fidélité avec l'original, et le fichier reste lisible/correct si
quelqu'un l'ouvre un jour directement dans Excel).
"""

from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Constantes DA Duchess
# --------------------------------------------------------------------------

GOLD = "E8DCC4"
GOLD_DARK = "C9B78C"
CREAM = "F8F5EE"
BLACK = "1A1A1A"
DEPENSE_GREY = "595959"
BORDER_GREY = "999999"
FONT_NAME = "Raleway"
EUR_FMT = '#,##0.00 "€"'
PCT_FMT = "0%"

COL_CAT = 1
COL_AXE = 2
COL_LABEL = 3
COL_FOURNISSEUR = 4
COL_BUDGET = 5
COL_PREVI = 6
COL_REALISE = 7
COL_PREVI_BUDGET = 8
COL_REALISE_PREVI = 9
COL_REALISE_BUDGET = 10
COL_TYPE = 11
N_COLS = 10

TYPE_CAT = "CAT"
TYPE_SP = "SP"
TYPE_DEP = "DEP"

HEADER_LAST_ROW = 12  # dernière ligne d'en-tête (colonnes A-K), les données commencent ligne 13
DATA_START_ROW = 13

CATEGORY_PRESETS: dict[str, list[str]] = {
    "PHONO": [
        "Rémunérations artistiques", "Ingénieur du son", "Réalisateur", "Mixe", "Studio",
        "Mastering", "Location matériel / instruments", "Bandes et fournitures",
        "Hébergement / transport", "Défraiements artistes",
    ],
    "CLIPS": ["Clip", "Visualizer", "Lyric video", "Teaser"],
    "ARTWORK": ["Cession droit image", "Shooting photo", "Graphisme", "Direction artistique"],
    "PROMOTION": [
        "Photos de presse", "Attaché(e) de presse", "Achat espace publicitaire",
        "Partenariats", "EPK", "VHR", "Prestataires (ingé son, etc.)", "Showcase / Live Sessions",
    ],
    "PROMOTION DIGITALE": [
        "Création de contenu", "Campagne digitale (Meta, TikTok)", "Campagne off-line",
        "Community management", "Influence",
    ],
    "FABRICATION": ["Pressage CD", "Pressage Vinyle", "SDRM", "Merchandising", "Distribution"],
    "DIVERS": ["Consulting", "Avocat", "Établissement des contrats", "Notaire"],
}
CATEGORIES: list[str] = list(CATEGORY_PRESETS.keys())
SIMPLE_CATEGORIES = {"DIVERS"}  # sous-postes sans dépenses enfants (saisie directe E/F/G)

_THIN_GREY = Side(style="thin", color=BORDER_GREY)
_BORDER_ALL = Border(left=_THIN_GREY, right=_THIN_GREY, top=_THIN_GREY, bottom=_THIN_GREY)


def _num(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _ratio(a: float, b: float) -> Optional[float]:
    if not b:
        return None
    return a / b


# --------------------------------------------------------------------------
# En-tête visuel (logo + titre + bandeau or) — identique à la maquette validée
# --------------------------------------------------------------------------


def _band(ws: Worksheet, row: int, color: str, height: float = 20):
    ws.row_dimensions[row].height = height
    for c in range(1, N_COLS + 2):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=color)


def build_header(ws: Worksheet, artist: str, project_label: str, logo_path: Optional[str] = None):
    ws.sheet_view.showGridLines = False
    for row, h in [(1, 8), (2, 40), (3, 26), (4, 6), (5, 30), (6, 14)]:
        _band(ws, row, BLACK, h)

    if logo_path:
        try:
            from openpyxl.drawing.image import Image as XLImage

            img = XLImage(logo_path)
            target_w = 62
            img.width = target_w
            img.height = int(img.height * (target_w / img.width)) if img.width else 40
            ws.add_image(img, "A2")
        except Exception:  # noqa: BLE001 — jamais bloquant, le fichier reste valable sans logo
            pass

    ws["C3"] = "BUDGET DE PRODUCTION"
    ws["C3"].font = Font(name=FONT_NAME, size=18, bold=True, color=GOLD)

    ws["C5"] = "ARTISTE"
    ws["C5"].font = Font(name=FONT_NAME, size=8.5, bold=True, color=GOLD_DARK)
    ws["D5"] = artist
    ws["D5"].font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")

    ws["G5"] = "PROJET"
    ws["G5"].font = Font(name=FONT_NAME, size=8.5, bold=True, color=GOLD_DARK)
    ws["H5"] = project_label
    ws["H5"].font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")

    _band(ws, 7, GOLD_DARK, 4)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    for col in "EFG":
        ws.column_dimensions[col].width = 15
    for col in "HIJ":
        ws.column_dimensions[col].width = 10
    ws.column_dimensions["K"].hidden = True


def build_column_headers(ws: Worksheet):
    labels_row10 = {5: "BUDGET", 6: "PRÉVISIONNEL", 7: "RÉALISÉ"}
    ratio_headers = {
        8: ("Prévi /\nBudget", "conso budget"),
        9: ("Réalisé /\nPrévi", "conso prévi"),
        10: ("Réalisé /\nBudget", "conso budget réel"),
    }
    ws.row_dimensions[10].height = 28
    ws.row_dimensions[11].height = 14
    for col, text in labels_row10.items():
        cell = ws.cell(row=10, column=col, value=text)
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, (main, sub) in ratio_headers.items():
        cell = ws.cell(row=10, column=col, value=main)
        cell.font = Font(name=FONT_NAME, size=8.5, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sub_cell = ws.cell(row=11, column=col, value=sub)
        sub_cell.font = Font(name=FONT_NAME, size=8, italic=True, color=DEPENSE_GREY)
        sub_cell.alignment = Alignment(horizontal="center")

    _band(ws, 12, BLACK, 20)
    headers = {1: "Catégorie", 2: "Axe analytique 3", 3: "Dépense", 4: "Fournisseur"}
    for col, text in headers.items():
        cell = ws.cell(row=12, column=col, value=text)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left" if col != 2 else "center", vertical="center")


def _cf_conso(ws: Worksheet, col_letter: str, start_row: int, end_row: int):
    # ISNUMBER(...) garde : IFERROR(...,"") renvoie "" quand budget/prévi = 0 (division impossible) —
    # sans ce garde-fou, LibreOffice/Excel comparent "" à un nombre et colorent ces cases vides en
    # rouge par erreur (constaté en test réel). On ne colore que les vraies valeurs numériques.
    ref = f"{col_letter}{start_row}"
    row_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=[f'AND(ISNUMBER({ref}),{ref}>1)'], fill=PatternFill("solid", fgColor="F4C7C3")),
    )
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=[f'AND(ISNUMBER({ref}),{ref}>=0.9,{ref}<=1)'], fill=PatternFill("solid", fgColor="FCE8B2")),
    )
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=[f'AND(ISNUMBER({ref}),{ref}<0.9)'], fill=PatternFill("solid", fgColor="C6E0B4")),
    )


# --------------------------------------------------------------------------
# Écriture des lignes (miroir des Sub Style_* / Inserer_* de la macro)
# --------------------------------------------------------------------------


def _clear_row(ws: Worksheet, r: int):
    for c in range(1, N_COLS + 2):
        ws.cell(row=r, column=c).value = None
        ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)


def _write_category_row(ws: Worksheet, r: int, name: str, budget: float, previ_formula: str, realise_formula: str):
    ws.row_dimensions[r].height = 22
    for c in range(1, N_COLS + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=GOLD)
        cell.font = Font(name=FONT_NAME, size=12, bold=True, color=BLACK)
        cell.border = _BORDER_ALL
    ws.cell(row=r, column=COL_CAT, value=name)
    ws.cell(row=r, column=COL_BUDGET, value=budget).number_format = EUR_FMT
    f_cell = ws.cell(row=r, column=COL_PREVI, value=previ_formula)
    f_cell.number_format = EUR_FMT
    g_cell = ws.cell(row=r, column=COL_REALISE, value=realise_formula)
    g_cell.number_format = EUR_FMT
    for col in (COL_BUDGET, COL_PREVI, COL_REALISE):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=r, column=COL_PREVI_BUDGET, value=f'=IFERROR(F{r}/E{r},"")').number_format = PCT_FMT
    ws.cell(row=r, column=COL_REALISE_PREVI, value=f'=IFERROR(G{r}/F{r},"")').number_format = PCT_FMT
    ws.cell(row=r, column=COL_REALISE_BUDGET, value=f'=IFERROR(G{r}/E{r},"")').number_format = PCT_FMT
    for col in (COL_PREVI_BUDGET, COL_REALISE_PREVI, COL_REALISE_BUDGET):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=COL_TYPE, value=TYPE_CAT)


def _write_container_row(ws: Worksheet, r: int, name: str, budget: float, dep_first: int, dep_last: int):
    ws.row_dimensions[r].height = 22
    for c in range(1, N_COLS + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=GOLD)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=BLACK)
        cell.border = _BORDER_ALL
    label_cell = ws.cell(row=r, column=COL_LABEL, value=f"{name} :")
    label_cell.alignment = Alignment(horizontal="left", indent=2)
    ws.cell(row=r, column=COL_BUDGET, value=budget).number_format = EUR_FMT
    if dep_last >= dep_first:
        ws.cell(row=r, column=COL_PREVI, value=f"=SUM(F{dep_first}:F{dep_last})").number_format = EUR_FMT
        ws.cell(row=r, column=COL_REALISE, value=f"=SUM(G{dep_first}:G{dep_last})").number_format = EUR_FMT
    else:
        ws.cell(row=r, column=COL_PREVI, value=0).number_format = EUR_FMT
        ws.cell(row=r, column=COL_REALISE, value=0).number_format = EUR_FMT
    for col in (COL_BUDGET, COL_PREVI, COL_REALISE):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=r, column=COL_PREVI_BUDGET, value=f'=IFERROR(F{r}/E{r},"")').number_format = PCT_FMT
    ws.cell(row=r, column=COL_REALISE_PREVI, value=f'=IFERROR(G{r}/F{r},"")').number_format = PCT_FMT
    ws.cell(row=r, column=COL_REALISE_BUDGET, value=f'=IFERROR(G{r}/E{r},"")').number_format = PCT_FMT
    for col in (COL_PREVI_BUDGET, COL_REALISE_PREVI, COL_REALISE_BUDGET):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=COL_TYPE, value=TYPE_SP)


def _write_depense_row(ws: Worksheet, r: int, label: str, fournisseur: str, previ: float, realise: float):
    ws.row_dimensions[r].height = 18
    for c in range(1, N_COLS + 1):
        ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)
        ws.cell(row=r, column=c).border = _BORDER_ALL
    lbl = ws.cell(row=r, column=COL_LABEL, value=label)
    lbl.font = Font(name=FONT_NAME, size=10, italic=True, color=DEPENSE_GREY)
    lbl.alignment = Alignment(horizontal="left", indent=4)
    f_cell = ws.cell(row=r, column=COL_FOURNISSEUR, value=fournisseur or "")
    f_cell.fill = PatternFill("solid", fgColor=CREAM)
    f_cell.font = Font(name=FONT_NAME, size=10, italic=True, color=DEPENSE_GREY)
    for col, val in ((COL_PREVI, previ), (COL_REALISE, realise)):
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.font = Font(name=FONT_NAME, size=10, color=BLACK)
        cell.number_format = EUR_FMT
        cell.alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=r, column=COL_TYPE, value=TYPE_DEP)


def _write_simple_row(ws: Worksheet, r: int, name: str, fournisseur: str, budget: float, previ: float, realise: float):
    ws.row_dimensions[r].height = 18
    for c in range(1, N_COLS + 1):
        ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=11)
        ws.cell(row=r, column=c).border = _BORDER_ALL
    lbl = ws.cell(row=r, column=COL_LABEL, value=name)
    lbl.alignment = Alignment(horizontal="left", indent=2)
    f_cell = ws.cell(row=r, column=COL_FOURNISSEUR, value=fournisseur or "")
    f_cell.fill = PatternFill("solid", fgColor=CREAM)
    for col, val in ((COL_BUDGET, budget), (COL_PREVI, previ), (COL_REALISE, realise)):
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.font = Font(name=FONT_NAME, size=10, color=BLACK)
        cell.number_format = EUR_FMT
        cell.alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=r, column=COL_PREVI_BUDGET, value=f'=IFERROR(F{r}/E{r},"")').number_format = PCT_FMT
    ws.cell(row=r, column=COL_REALISE_PREVI, value=f'=IFERROR(G{r}/F{r},"")').number_format = PCT_FMT
    ws.cell(row=r, column=COL_REALISE_BUDGET, value=f'=IFERROR(G{r}/E{r},"")').number_format = PCT_FMT
    for col in (COL_PREVI_BUDGET, COL_REALISE_PREVI, COL_REALISE_BUDGET):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=COL_TYPE, value=TYPE_SP)


def _write_total_row(ws: Worksheet, r: int, cat_rows: list[int]):
    ws.row_dimensions[r].height = 24
    for c in range(1, N_COLS + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=BLACK)
        cell.font = Font(name=FONT_NAME, size=12, bold=True, color=GOLD)
        cell.border = _BORDER_ALL
    ws.cell(row=r, column=COL_LABEL, value="TOTAL DÉPENSES")
    if cat_rows:
        e_sum = "+".join(f"E{cr}" for cr in cat_rows)
        f_sum = "+".join(f"F{cr}" for cr in cat_rows)
        g_sum = "+".join(f"G{cr}" for cr in cat_rows)
    else:
        e_sum = f_sum = g_sum = "0"
    ws.cell(row=r, column=COL_BUDGET, value=f"={e_sum}").number_format = EUR_FMT
    ws.cell(row=r, column=COL_PREVI, value=f"={f_sum}").number_format = EUR_FMT
    ws.cell(row=r, column=COL_REALISE, value=f"={g_sum}").number_format = EUR_FMT
    for col in (COL_BUDGET, COL_PREVI, COL_REALISE):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="right", indent=1)


# --------------------------------------------------------------------------
# Rendu complet d'un feuillet à partir de l'arbre JSON
# --------------------------------------------------------------------------


def render_tree_to_sheet(ws: Worksheet, artist: str, project_label: str, tree: dict, logo_path: Optional[str] = None):
    """Efface tout le contenu de données existant (>= ligne 1) et redessine
    entièrement le feuillet (en-tête + colonnes + arbre catégories) à partir
    de `tree` (voir docstring du module pour la forme attendue)."""
    max_row = max(ws.max_row, 200)
    for r in range(1, max_row + 1):
        _clear_row(ws, r)
    ws.conditional_formatting = type(ws.conditional_formatting)()

    build_header(ws, artist, project_label, logo_path)
    build_column_headers(ws)

    categories_in = {c.get("name"): c for c in (tree.get("categories") or [])}
    r = DATA_START_ROW
    cat_rows: list[int] = []

    for cat_name in CATEGORIES:
        cat = categories_in.get(cat_name, {"name": cat_name, "budget": 0, "sous_postes": []})
        cat_row = r
        cat_rows.append(cat_row)
        sp_rows: list[int] = []
        r += 1
        for sp in cat.get("sous_postes", []) or []:
            sp_type = sp.get("type") or ("simple" if cat_name in SIMPLE_CATEGORIES else "container")
            if sp_type == "simple":
                sp_rows.append(r)
                _write_simple_row(
                    ws, r, sp.get("name", ""), sp.get("fournisseur", ""),
                    _num(sp.get("budget")), _num(sp.get("previsionnel")), _num(sp.get("realise")),
                )
                r += 1
            else:
                container_row = r
                sp_rows.append(container_row)
                r += 1
                dep_first = r
                depenses = sp.get("depenses") or []
                if not depenses:
                    depenses = [{"label": "Dépense n°1", "fournisseur": "", "previsionnel": 0, "realise": 0}]
                for dep in depenses:
                    _write_depense_row(
                        ws, r, dep.get("label") or f"Dépense n°{r - dep_first + 1}",
                        dep.get("fournisseur", ""), _num(dep.get("previsionnel")), _num(dep.get("realise")),
                    )
                    r += 1
                dep_last = r - 1
                _write_container_row(ws, container_row, sp.get("name", ""), _num(sp.get("budget")), dep_first, dep_last)
            # ligne vide de séparation entre sous-postes
            ws.row_dimensions[r].height = 8
            r += 1
        if sp_rows:
            p1, fin = sp_rows[0], sp_rows[-1]
            # fin de bloc = juste avant la ligne catégorie suivante (r actuel, avant incrément)
            bloc_fin = r - 1
            previ_formula = f'=SUMIFS(F{p1}:F{bloc_fin},K{p1}:K{bloc_fin},"{TYPE_SP}")'
            realise_formula = f'=SUMIFS(G{p1}:G{bloc_fin},K{p1}:K{bloc_fin},"{TYPE_SP}")'
        else:
            previ_formula, realise_formula = 0, 0
        _write_category_row(ws, cat_row, cat_name, _num(cat.get("budget")), previ_formula, realise_formula)
        # ligne vide entre catégories
        ws.row_dimensions[r].height = 10
        r += 1

    total_row = r
    _write_total_row(ws, total_row, cat_rows)

    for col_letter in ("H", "I", "J"):
        _cf_conso(ws, col_letter, DATA_START_ROW, total_row)

    ws.print_area = f"A1:K{total_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return total_row


def new_project_sheet(wb: Workbook, sheet_name: str, artist: str, project_label: str, logo_path: Optional[str] = None) -> Worksheet:
    ws = wb.create_sheet(title=sheet_name[:31])
    empty_tree = {"categories": [{"name": c, "budget": 0, "sous_postes": []} for c in CATEGORIES]}
    render_tree_to_sheet(ws, artist, project_label, empty_tree, logo_path)
    return ws


def new_artist_workbook(artist: str, project_label: str, logo_path: Optional[str] = None) -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    new_project_sheet(wb, project_label[:31] or "Projet 1", artist, project_label, logo_path)
    return wb


# --------------------------------------------------------------------------
# Lecture d'un feuillet existant -> arbre JSON (+ totaux/ratios recalculés)
# --------------------------------------------------------------------------


def parse_sheet_to_tree(ws: Worksheet) -> dict:
    categories: list[dict] = []
    current_cat: Optional[dict] = None
    current_sp: Optional[dict] = None

    max_row = ws.max_row
    for r in range(DATA_START_ROW, max_row + 1):
        row_type = ws.cell(row=r, column=COL_TYPE).value
        row_type = (str(row_type).strip() if row_type is not None else "")
        label = ws.cell(row=r, column=COL_LABEL).value
        cat_name_cell = ws.cell(row=r, column=COL_CAT).value

        if row_type == TYPE_CAT:
            name = (cat_name_cell or "").strip() if isinstance(cat_name_cell, str) else str(cat_name_cell or "")
            if name.upper() == "TOTAL DÉPENSES" or (label and str(label).strip().upper() == "TOTAL DÉPENSES"):
                continue
            current_cat = {
                "name": name,
                "budget": _num(ws.cell(row=r, column=COL_BUDGET).value),
                "sous_postes": [],
            }
            categories.append(current_cat)
            current_sp = None
        elif row_type == TYPE_SP:
            if current_cat is None:
                continue
            name_raw = str(label or "").strip()
            is_container = name_raw.endswith(":")
            sp_name = name_raw[:-1].strip() if is_container else name_raw
            if is_container:
                current_sp = {
                    "name": sp_name,
                    "type": "container",
                    "budget": _num(ws.cell(row=r, column=COL_BUDGET).value),
                    "depenses": [],
                }
            else:
                current_sp = {
                    "name": sp_name,
                    "type": "simple",
                    "fournisseur": ws.cell(row=r, column=COL_FOURNISSEUR).value or "",
                    "budget": _num(ws.cell(row=r, column=COL_BUDGET).value),
                    "previsionnel": _num(ws.cell(row=r, column=COL_PREVI).value),
                    "realise": _num(ws.cell(row=r, column=COL_REALISE).value),
                }
            current_cat["sous_postes"].append(current_sp)
        elif row_type == TYPE_DEP:
            if current_sp is None or current_sp.get("type") != "container":
                continue
            current_sp["depenses"].append({
                "label": label or "",
                "fournisseur": ws.cell(row=r, column=COL_FOURNISSEUR).value or "",
                "previsionnel": _num(ws.cell(row=r, column=COL_PREVI).value),
                "realise": _num(ws.cell(row=r, column=COL_REALISE).value),
            })

    return {"categories": categories}


def compute_totals(tree: dict) -> dict:
    """Recalcule (bottom-up, toujours en Python — jamais depuis les formules
    Excel, potentiellement pas recalculées) previsionnel/realise + les 3
    ratios à chaque niveau, et ajoute un total général."""
    grand = {"budget": 0.0, "previsionnel": 0.0, "realise": 0.0}
    for cat in tree.get("categories", []):
        cat_previ = cat_realise = 0.0
        for sp in cat.get("sous_postes", []):
            if sp.get("type") == "container":
                sp_previ = sum(_num(d.get("previsionnel")) for d in sp.get("depenses", []))
                sp_realise = sum(_num(d.get("realise")) for d in sp.get("depenses", []))
                sp["previsionnel"] = sp_previ
                sp["realise"] = sp_realise
            else:
                sp_previ = _num(sp.get("previsionnel"))
                sp_realise = _num(sp.get("realise"))
            sp_budget = _num(sp.get("budget"))
            sp["ratios"] = {
                "previ_budget": _ratio(sp_previ, sp_budget),
                "realise_previ": _ratio(sp_realise, sp_previ),
                "realise_budget": _ratio(sp_realise, sp_budget),
            }
            cat_previ += sp_previ
            cat_realise += sp_realise
        cat["previsionnel"] = cat_previ
        cat["realise"] = cat_realise
        cat_budget = _num(cat.get("budget"))
        cat["ratios"] = {
            "previ_budget": _ratio(cat_previ, cat_budget),
            "realise_previ": _ratio(cat_realise, cat_previ),
            "realise_budget": _ratio(cat_realise, cat_budget),
        }
        grand["budget"] += cat_budget
        grand["previsionnel"] += cat_previ
        grand["realise"] += cat_realise
    grand["ratios"] = {
        "previ_budget": _ratio(grand["previsionnel"], grand["budget"]),
        "realise_previ": _ratio(grand["realise"], grand["previsionnel"]),
        "realise_budget": _ratio(grand["realise"], grand["budget"]),
    }
    tree["total"] = grand
    return tree


# --------------------------------------------------------------------------
# Helpers haut niveau (bytes <-> tree), utilisés directement par le backend
# --------------------------------------------------------------------------


def workbook_from_bytes(data: bytes) -> Workbook:
    return load_workbook(io.BytesIO(data))


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_project_tree(data: bytes, sheet_name: str) -> dict:
    wb = workbook_from_bytes(data)
    if sheet_name not in wb.sheetnames:
        raise KeyError(sheet_name)
    tree = parse_sheet_to_tree(wb[sheet_name])
    return compute_totals(tree)


def write_project_tree(data: bytes, sheet_name: str, artist: str, project_label: str, tree: dict, logo_path: Optional[str] = None) -> bytes:
    wb = workbook_from_bytes(data)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name[:31])
    render_tree_to_sheet(ws, artist, project_label, tree, logo_path)
    return workbook_to_bytes(wb)
